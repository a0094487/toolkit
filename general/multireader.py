#reads pdf, docx, xlsx and any other files,
#successful reads outputs datastructure in the form of a dictionary with filename,
#unsuccessful appended into seperate list
#Choices of file, files in directory one level, and os.walk files
#.doc and .xlx not supported yet unfortunately, lookup panda export .xlx to .csv...
import PyPDF2, docx, csv, openpyxl, re, os
from openpyxl.utils import get_column_letter, column_index_from_string

def fileread(filename):
    if filename.endswith('.docx'):
        doc = docx.Document(filename)
        fullText = []
        for para in doc.paragraphs: fullText.append(para.text) #list of paragraphs
        dataobject[filename] = ('\n'.join(fullText)) #one long string of all paragraphs seperated by linebreaks, \t for tabs, etc
        
    elif filename.endswith('.pdf'):
        pdfFileObj = open(filename, 'rb') #read binary mode
        pdfReader = PyPDF2.PdfFileReader(pdfFileObj) #parse
        if not pdfReader.isEncrypted:
            numberofpages = pdfReader.numPages
            pg = 0
            dataobject[filename] = {}
            while pg < numberofpages:
                pageObj = pdfReader.getPage(pg)
                pg = pg + 1
                dataobject[filename][pg]=pageObj.extractText()
        pdfFileObj.close()

    elif filename.endswith('.csv'):
        csvRows = []
        csvFileObj = open(filename)
        readerObj = csv.reader(csvFileObj)
        #csvRows = list(readerObj) #less details
        for row in readerObj: 
        #    csvRows.append(row.extend(str(readerObj.line_num))) #with line_num appended in end of row.
            csvRows.append(row.insert(0, str(readerObj.line_num))) #alternatively, inserts .line_num as row[0] 
        dataobject[filename] = csvRows
        csvFileObj.close()

    elif filename.endswith('.xlsx') or filename.endswith('.xlsm') or filename.endswith('.xltx') or filename.endswith('.xltm'):
        dataobject[filename]={}
        #workbook = openpyxl.load_workbook(filename) #saves as values as original formulas
        workbook = openpyxl.load_workbook(filename, data_only=True)
        worksheets = workbook.get_sheet_names()
        for worksheet in worksheets: 
            sheet=workbook.get_sheet_by_name(worksheet)
            dataobject[filename][sheet.title]={}
            hRow=sheet.max_row #limitation for max_: uninterupted space from a1, cant have blanks
            hColumn=sheet.max_column
            for i in range(1,hRow,1):
                for j in range(1,hColumn,1):
                    dataobject[filename][sheet.title][(j,i)]=sheet.cell(row=i, column=j).value # "(x,y)" tuple reference, swap to (y,x) as needed.
                    #columnLetter=openpyxl.utils.get_column_letter(j) #inverse is column_index_from_string()
                    #cellLetter= columnLetter + str(i)
                    #dataobject[filename][sheet.title][cellLetter]=sheet.cell(row=i, column=j).value
                    #above 3 lines record in excel reference, swap in as needed.
                    #note: date/datatime/time objects retained. use .strftime('%Y/%m/%d'/%H/%m/%s) seperately to return string.          
        
    else:
        #continue 
        try: 
            openfile = open(filename)
            readFile = openfile.readlines()
            lines = []
            for linestring in readFile: lines.append(linestring)
            openfile.close()
            dataobject[filename] = lines
        except: 
            #print('Unable to read', filename ) #gives up reading.
            failedread.append(filename)
    return dataobject

dataobject = {} #keep this line of code above filewalk or dir loop.
failedread = []

print('In which directory do you want to search? Provide input of the form: C:\dir1\subdir1\subdir2')
userdir = str(input())
dirfiles = os.listdir(userdir)

##os.walk method; loops for all files and subdirectories
for dirpath, subdirs, files in os.walk(userdir):
    for filename in files:
        if os.path.isfile(str(dirpath) + "\\" + filename):
            filename = os.path.join(os.path.abspath(dirpath), filename)
            try:
                fileread(filename)
            except: continue

##Folder only option; loops only for all files, no subdirectories
#for i in range(len(dirfiles)):
#    filename = dirfiles[i]
#    readdpath = re.compile(r'[a-zA-Z]\:\\') #re of something like "C:\\"
#    if not readdpath.search(filename): filename = os.path.join(os.getcwd(), filename) #fills in full path into filename if user lazy
#    try: fileread(filename)
#    except: continue

##Specific file option; no file looping
#print('Which file do you want to read? Provide input of the form: C:\\dir1\\subdir1\\file1')
#filename = str(input())
#readdpath = re.compile(r'[a-zA-Z]\:\\') #re of something like "C:\\"
#if not readdpath.search(filename): filename = os.path.join(os.getcwd(), filename)
#fileread(filename)

#print(dataobject)
#print(failedread)
